import os
import re
import json
import time
import random
import sqlite3
import hashlib
import requests
import asyncio
import logging
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urlunparse

# Selenium & Webdriver Imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from fake_useragent import UserAgent

# Telegram Bot Imports
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ContextTypes,
    filters
)

# Logging Setup
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ==============================================================================
# CLASS: GoogleDorkerPro (المحرك الرئيسي المطور)
# ==============================================================================

class GoogleDorkerPro:
    def __init__(self, headless=True, ultra_slow=False):
        self.driver = None
        self.found_urls = set()
        self.stats_file = "dork_stats.json"
        self.ua = UserAgent()
        self.current_dork = ""
        self.current_page = 1
        self.dork_results = {}
        self.start_time = None
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.MAX_PAGES = 50
        self.total_dorks_processed = 0
        self.captcha_count = 0
        self.processed_dorks = []
        self.current_dork_index = 0
        self.headless = headless
        self.ultra_slow = ultra_slow
        self.delay_between_actions = 1 if not ultra_slow else 3
        self.emails_found = set()
        
        self.blacklist_domains = set()
        self.whitelist_domains = set()
        
        os.makedirs("results", exist_ok=True)
        os.makedirs("sessions", exist_ok=True)
        os.makedirs("emails", exist_ok=True)
        
        self.output_file = f"results/{self.session_id}_results.txt"
        self.emails_file = f"emails/emails_{self.session_id}.txt"
        
        self.load_existing_urls()
        self.load_stats()
        self.load_filters()
        self.init_database()
    
    def load_existing_urls(self):
        results_dir = "results"
        if os.path.exists(results_dir):
            for file in os.listdir(results_dir):
                if file.endswith("_results.txt"):
                    filepath = os.path.join(results_dir, file)
                    try:
                        with open(filepath, 'r', encoding='utf-8') as f:
                            for line in f:
                                match = re.search(r'https?://[^\s]+', line)
                                if match:
                                    self.found_urls.add(match.group())
                    except Exception:
                        pass
        print(f"📂 [SYSTEM] Loaded {len(self.found_urls)} existing URLs.")
    
    def load_stats(self):
        if os.path.exists(self.stats_file):
            try:
                with open(self.stats_file, 'r', encoding='utf-8') as f:
                    self.dork_results = json.load(f)
            except Exception:
                self.dork_results = {}
    
    def save_stats(self):
        with open(self.stats_file, 'w', encoding='utf-8') as f:
            json.dump(self.dork_results, f, ensure_ascii=False, indent=2)
    
    def load_filters(self):
        if os.path.exists("blacklist.txt"):
            with open("blacklist.txt", 'r', encoding='utf-8') as f:
                self.blacklist_domains.update(line.strip() for line in f if line.strip())
            print(f"🚫 [FILTER] Loaded {len(self.blacklist_domains)} blacklisted domains.")
        
        if os.path.exists("whitelist.txt"):
            with open("whitelist.txt", 'r', encoding='utf-8') as f:
                self.whitelist_domains.update(line.strip() for line in f if line.strip())
            print(f"✅ [FILTER] Loaded {len(self.whitelist_domains)} whitelisted domains.")
    
    def init_database(self):
        self.db_file = f"results/dorker_{self.session_id}.db"
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS urls (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE,
                domain TEXT,
                timestamp TEXT,
                dork TEXT,
                status_code INTEGER,
                is_alive BOOLEAN
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS dorks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dork TEXT,
                urls_count INTEGER,
                timestamp TEXT
            )
        ''')
        conn.commit()
        conn.close()
        print(f"💾 [DATABASE] Initialized: {self.db_file}")
    
    def save_to_database(self, url, dork="", status_code=None, is_alive=None):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            domain = urlparse(url).netloc
            cursor.execute('''
                INSERT OR REPLACE INTO urls (url, domain, timestamp, dork, status_code, is_alive)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (url, domain, datetime.now().isoformat(), dork, status_code, is_alive))
            conn.commit()
            conn.close()
        except Exception:
            pass
    
    def save_session(self):
        session_data = {
            'found_urls': list(self.found_urls),
            'processed_dorks': self.processed_dorks,
            'current_dork_index': self.current_dork_index,
            'session_id': self.session_id,
            'timestamp': datetime.now().isoformat(),
            'total_urls': len(self.found_urls)
        }
        session_file = f"sessions/session_{self.session_id}.json"
        with open(session_file, 'w', encoding='utf-8') as f:
            json.dump(session_data, f, ensure_ascii=False, indent=2)
        print(f"💾 [SESSION] Saved: {session_file}")
    
    def check_url_status(self, url):
        try:
            headers = {'User-Agent': self.ua.random}
            response = requests.head(url, headers=headers, timeout=5, allow_redirects=True)
            return url, response.status_code, response.status_code == 200
        except Exception:
            return url, 0, False
    
    def check_all_urls(self):
        print("\n🔍 [CHECKER] Verifying URL accessibility...")
        alive_urls = []
        dead_urls = []
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = {executor.submit(self.check_url_status, url): url for url in self.found_urls}
            
            for i, future in enumerate(as_completed(futures), 1):
                url, status, is_alive = future.result()
                self.save_to_database(url, status_code=status, is_alive=is_alive)
                if is_alive:
                    alive_urls.append(url)
                else:
                    dead_urls.append(url)
        
        with open(f"results/alive_{self.session_id}.txt", 'w', encoding='utf-8') as f:
            for url in alive_urls:
                f.write(f"{url}\n")
        
        return alive_urls, dead_urls
    
    def extract_emails_from_page(self):
        try:
            page_text = self.driver.page_source
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            emails = set(re.findall(email_pattern, page_text))
            
            exclude = ['example.com', 'test.com', 'noreply', 'no-reply', 'support', 'admin', 'sentry.io', 'domain.com']
            emails = [e for e in emails if not any(x in e.lower() for x in exclude)]
            
            with open(self.emails_file, 'a', encoding='utf-8') as f:
                for email in emails:
                    if email not in self.emails_found:
                        f.write(f"{email}\n")
                        self.emails_found.add(email)
            
            return emails
        except Exception:
            return set()
    
    def export_to_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dork Results"
            
            headers = ['#', 'URL', 'Domain', 'Timestamp']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for i, url in enumerate(self.found_urls, 1):
                ws.cell(row=i+1, column=1, value=i)
                ws.cell(row=i+1, column=2, value=url)
                ws.cell(row=i+1, column=3, value=urlparse(url).netloc)
                ws.cell(row=i+1, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 30
            
            excel_file = f"results/dork_results_{self.session_id}.xlsx"
            wb.save(excel_file)
            return excel_file
        except ImportError:
            return None
    
    def generate_advanced_report(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            pdf_file = f"results/report_{self.session_id}.pdf"
            doc = SimpleDocTemplate(pdf_file, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#1F4E79'))
            story.append(Paragraph("Google Dorker Pro - Intelligence Report", title_style))
            story.append(Spacer(1, 0.2*inch))
            
            duration = datetime.now() - self.start_time if self.start_time else datetime.now() - datetime.now()
            seconds = int(duration.total_seconds())
            
            stats_data = [
                ['Metric', 'Value'],
                ['Total URLs Found', str(len(self.found_urls))],
                ['Dorks Processed', str(self.total_dorks_processed)],
                ['Pages Scanned', str(self.current_page)],
                ['Captchas Enriched', str(self.captcha_count)],
                ['Emails Extracted', str(len(self.emails_found))],
                ['Duration', f"{seconds // 60}m {seconds % 60}s"],
            ]
            
            table = Table(stats_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            story.append(table)
            
            doc.build(story)
            return pdf_file
        except Exception:
            return None
    
    def create_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--headless=new")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--lang=en-US,en")
        chrome_options.add_argument("--remote-debugging-port=9222")
        chrome_options.add_argument("--disable-setuid-sandbox")
        
        for bin_path in ["/usr/bin/chromium", "/usr/bin/chromium-browser", "/usr/bin/google-chrome"]:
            if os.path.exists(bin_path):
                chrome_options.binary_location = bin_path
                break

        try:
            random_ua = self.ua.random
            chrome_options.add_argument(f"user-agent={random_ua}")
        except Exception:
            chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

        driver_path = None
        for path in ["/usr/bin/chromedriver", "/usr/lib/chromium-browser/chromedriver"]:
            if os.path.exists(path):
                driver_path = path
                break

        if driver_path:
            service = Service(executable_path=driver_path)
        else:
            service = Service(ChromeDriverManager().install())
            
        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.driver.set_page_load_timeout(30)
        return self.driver
    
    def clean_url(self, url):
        try:
            parsed = urlparse(url)
            clean = urlunparse((parsed.scheme, parsed.netloc, parsed.path, '', '', ''))
            if len(clean) > 250:
                return None
            return clean
        except Exception:
            return None
    
    def is_allowed_url(self, url):
        domain = urlparse(url).netloc
        if domain in self.blacklist_domains:
            return False
        if self.whitelist_domains and domain not in self.whitelist_domains:
            return False
        return True
    
    def save_url(self, url, source_dork=""):
        if not self.is_allowed_url(url):
            return False
        
        with open(self.output_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            dork_hash = hashlib.md5(source_dork.encode()).hexdigest()[:8] if source_dork else "unknown"
            f.write(f"[{timestamp}] [{dork_hash}] {url}\n")
        
        self.found_urls.add(url)
        self.save_to_database(url, dork=source_dork)
        return True
    
    def save_dork_results(self, dork, urls_count):
        dork_hash = hashlib.md5(dork.encode()).hexdigest()
        self.dork_results[dork_hash] = {
            'dork_preview': dork[:100],
            'urls_found': urls_count,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'pages_searched': self.current_page,
            'session_id': self.session_id
        }
        self.save_stats()
    
    def detect_captcha(self):
        try:
            captcha_selectors = [
                "//iframe[contains(@src, 'recaptcha')]",
                "//iframe[contains(@src, 'captcha')]",
                "//div[contains(@class, 'g-recaptcha')]",
                "//*[contains(text(), 'unusual traffic')]",
                "//*[contains(text(), 'verify you are human')]",
            ]
            for selector in captcha_selectors:
                if self.driver.find_elements(By.XPATH, selector):
                    self.captcha_count += 1
                    return True
            return False
        except Exception:
            return False
    
    def handle_captcha(self):
        if self.detect_captcha():
            print(f"⚠️ [CAPTCHA] Detected CAPTCHA #{self.captcha_count}")
            time.sleep(3)
            return self.restore_search()
        return True
    
    def restore_search(self):
        try:
            if "google.com/search" not in self.driver.current_url:
                self.driver.get("https://www.google.com")
                time.sleep(2)
                
                search_box = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.NAME, "q"))
                )
                search_box.clear()
                for char in self.current_dork:
                    search_box.send_keys(char)
                    time.sleep(random.uniform(0.005, 0.02))
                search_box.send_keys(Keys.RETURN)
                time.sleep(3)
            return True
        except Exception as e:
            print(f"⚠️ [RESTORE ERROR] {str(e)}")
            return False
    
    def get_next_page_url(self):
        try:
            next_selectors = [
                "//a[@aria-label='Next page']",
                "//a[@aria-label='الصفحة التالية']",
                "//a[contains(@id, 'pnnext')]",
                "//span[text()='Next']/parent::a"
            ]
            for selector in next_selectors:
                elements = self.driver.find_elements(By.XPATH, selector)
                for element in elements:
                    if element.is_enabled():
                        href = element.get_attribute("href")
                        if href:
                            return href
            
            current_url = self.driver.current_url
            if 'start=' in current_url:
                return re.sub(r'start=\d+', f'start={self.current_page * 10}', current_url)
            else:
                return current_url + f'&start={self.current_page * 10}'
        except Exception:
            return None
    
    def extract_links(self):
        links = set()
        selectors = [
            "a[jsname='UWckNb']",
            "a[href^='http']:not([href*='google']):not([href*='youtube'])",
            "div#search a[href^='http']",
            "div.g a[href^='http']",
            "div.yuRUbf a",
        ]
        
        for selector in selectors:
            try:
                elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                for elem in elements:
                    try:
                        href = elem.get_attribute("href")
                        if href and href.startswith("http"):
                            exclude_list = ['google.com', 'youtube.com', 'support.google', 'accounts.google', 'gstatic.com']
                            if not any(x in href for x in exclude_list):
                                clean = self.clean_url(href)
                                if clean:
                                    links.add(clean)
                    except Exception:
                        continue
            except Exception:
                continue
        return links
    
    def show_statistics(self):
        duration = datetime.now() - self.start_time if self.start_time else datetime.now() - datetime.now()
        seconds = int(duration.total_seconds())
        
        stats_msg = "🔥 [GOOGLE DORKER PRO - ENGINE STATS] 🔥\n"
        stats_msg += f"----------------------------------------\n"
        stats_msg += f"✅ Total URLs Collected : {len(self.found_urls)}\n"
        stats_msg += f"📧 Total Emails Found   : {len(self.emails_found)}\n"
        stats_msg += f"⏱️ Active Time         : {seconds // 60}m {seconds % 60}s\n"
        stats_msg += f"📄 Pages Processed     : {self.current_page}\n"
        stats_msg += f"🤖 Captchas Enriched   : {self.captcha_count}\n"
        stats_msg += f"🎯 Dorks Executed      : {self.total_dorks_processed}\n"
        return stats_msg
    
    def search(self, dork):
        self.current_dork = dork
        self.current_page = 1
        if self.start_time is None:
            self.start_time = datetime.now()
        self.total_dorks_processed += 1
        self.processed_dorks.append(dork)
        
        if not self.driver:
            self.create_driver()
        
        print(f"\n🔍 [EXEC] Searching #{self.total_dorks_processed}: {dork}")
        
        try:
            self.driver.get("https://www.google.com")
            time.sleep(random.uniform(2, 4) * self.delay_between_actions)
            
            try:
                cookie_btn = WebDriverWait(self.driver, 4).until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(.,'Accept all') or contains(.,'Accept')]"))
                )
                cookie_btn.click()
            except Exception:
                pass
            
            search_box = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.NAME, "q"))
            )
            
            search_box.clear()
            for char in dork:
                search_box.send_keys(char)
                time.sleep(random.uniform(0.005, 0.02) * self.delay_between_actions)
            
            search_box.send_keys(Keys.RETURN)
            time.sleep(random.uniform(3, 5) * self.delay_between_actions)
            
            if not self.handle_captcha():
                return 0
            
            pages_without_new = 0
            urls_before = len(self.found_urls)
            
            while pages_without_new < 3 and self.current_page <= self.MAX_PAGES:
                try:
                    WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "div#search, div.g"))
                    )
                except Exception:
                    time.sleep(3)
                
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.5 * self.delay_between_actions)
                
                self.extract_emails_from_page()
                new_links = self.extract_links()
                
                new_count = 0
                for link in new_links:
                    if link not in self.found_urls:
                        if self.save_url(link, dork):
                            new_count += 1
                
                if new_count == 0:
                    pages_without_new += 1
                else:
                    pages_without_new = 0
                
                next_url = self.get_next_page_url()
                if not next_url:
                    break
                
                self.driver.get(next_url)
                time.sleep(random.uniform(3, 6) * self.delay_between_actions)
                
                if not self.handle_captcha():
                    break
                
                self.current_page += 1
            
            urls_found = len(self.found_urls) - urls_before
            self.save_dork_results(dork, urls_found)
            self.save_session()
            return urls_found
            
        except Exception as e:
            print(f"⚠️ [ENGINE ERROR] {str(e)}")
            return 0
    
    def export_all_results(self):
        excel_path = self.export_to_excel()
        pdf_path = self.generate_advanced_report()
        
        txt_path = f"results/urls_only_{self.session_id}.txt"
        with open(txt_path, 'w', encoding='utf-8') as f:
            for url in sorted(self.found_urls):
                f.write(f"{url}\n")
        return txt_path, excel_path, pdf_path
    
    def close(self):
        if self.driver:
            self.save_session()
            files = self.export_all_results()
            self.driver.quit()
            self.driver = None
            return files
        return None, None, None

# ==============================================================================
# TELEGRAM BOT CONTROLLER
# ==============================================================================

BOT_TOKEN = "8692960014:AAEpYPo0XTj8F2DmAeUgdaf9_w06MWFYDeI"

user_states = {}
dorker_instances = {}
user_files_data = {}

def get_main_keyboard():
    keyboard = [
        [
            InlineKeyboardButton("🎯 بحث دورك فردي (/dork)", callback_data="mode_single"),
            InlineKeyboardButton("📂 رفع ملف دوركات (.txt)", callback_data="mode_multi")
        ],
        [
            InlineKeyboardButton("📊 عرض الإحصائيات", callback_data="show_stats"),
            InlineKeyboardButton("⚡ فحص الروابط (Alive/Dead)", callback_data="check_urls")
        ],
        [
            InlineKeyboardButton("📦 تصدير وتحميل الحصيلة", callback_data="export_files"),
            InlineKeyboardButton("🛑 إيقاف وحفظ الجلسة", callback_data="stop_engine")
        ]
    ]
    return InlineKeyboardMarkup(keyboard)

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome_text = (
        "🔥 *GOOGLE DORKER PRO - SHADOW EDITION v3.5* 🔥\n\n"
        "⚡ *المحرك الذكي الأوتوماتيكي لاستخراج الروابط والبيانات*\n\n"
        "🛠️ *الأوامر المتاحة:*\n"
        "• `/dork <query>` - لبدء الفحص الفوري المباشر.\n"
        "• أرسل ملف `.txt` يحتوي قائمة دوركات لبدء لوحة الفحص الجماعية.\n"
        "• `/stats` - لعرض إحصائيات المحرك الحالية.\n"
        "• `/help` - دليل الاستخدام السريع."
    )
    if update.message:
        await update.message.reply_text(welcome_text, parse_mode="Markdown", reply_markup=get_main_keyboard())

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = (
        "📖 *دليل التشغيل الاحترافي:*\n\n"
        "1️⃣ **البحث الفردي المباشر:**\n"
        "`/dork inurl:admin.php`\n\n"
        "2️⃣ **رفع الملفات الجماعية:**\n"
        "قم بإرسال ملف نصي `.txt` يحتوي على الدوركات (كل دورك في سطر).\n\n"
        "3️⃣ **التصدير:**\n"
        "يمكنك سحب النتائج فوراً بصيغ Text / Excel / PDF."
    )
    if update.message:
        await update.message.reply_text(help_text, parse_mode="Markdown")

async def dork_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    if not context.args:
        if update.message:
            await update.message.reply_text("⚠️ يرجى تزويد الاستعلام بعد الأمر، مثال:\n`/dork inurl:login.php`", parse_mode="Markdown")
        return

    dork_text = " ".join(context.args)
    if update.message:
        await update.message.reply_text(f"🚀 *بدء محرك الفحص الفردي الناري:*\n`{dork_text}`", parse_mode="Markdown")

    if chat_id not in dorker_instances or dorker_instances[chat_id] is None:
        dorker_instances[chat_id] = GoogleDorkerPro(headless=True, ultra_slow=False)

    dorker = dorker_instances[chat_id]
    loop = asyncio.get_running_loop()
    added_now = await loop.run_in_executor(None, dorker.search, dork_text)

    if update.message:
        await update.message.reply_text(
            f"✅ *اكتمل فحص الدورك!*\n"
            f"➕ روابط مضافة حديثاً: `{added_now}`\n"
            f"🌐 إجمالي الروابط في القاعدة: `{len(dorker.found_urls)}`",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )

async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    target_msg = update.message if update.message else (update.callback_query.message if update.callback_query else None)
    
    if chat_id in dorker_instances and dorker_instances[chat_id]:
        stats = dorker_instances[chat_id].show_statistics()
        if target_msg:
            await target_msg.reply_text(f"```\n{stats}\n```", parse_mode="Markdown")
    else:
        if target_msg:
            await target_msg.reply_text("❌ لا يوجد محرك شغال حالياً. ابدأ بالبحث أولاً!")

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.document:
        return
        
    chat_id = update.effective_chat.id
    document = update.message.document

    if not document.file_name.endswith('.txt'):
        await update.message.reply_text("❌ يرجى رفع ملف نصي بتنسيق `.txt` فقط.")
        return

    await update.message.reply_text("📥 *جاري تحميل الملف وتصفية البيانات...*", parse_mode="Markdown")

    file = await context.bot.get_file(document.file_id)
    file_bytes = await file.download_as_bytearray()
    content = file_bytes.decode('utf-8', errors='ignore')

    raw_lines = [line.strip() for line in content.splitlines() if line.strip()]
    unique_lines = list(dict.fromkeys(raw_lines))

    user_files_data[chat_id] = unique_lines

    panel_text = (
        "🔥 *لوحة التحكم بالملف المرفوع* 🔥\n\n"
        f"📄 **اسم الملف:** `{document.file_name}`\n"
        f"🔢 **إجمالي الأسطر:** `{len(raw_lines)}`\n"
        f"✨ **الدوركات الصافية (بدون تكرار):** `{len(unique_lines)}`\n\n"
        "اختر الإجراء المطلوب لبدء معالجة الدوركات واستخراج الروابط:"
    )

    keyboard = [
        [
            InlineKeyboardButton("⚡ بدء الفحص الناري فوراً", callback_data="start_file_process"),
            InlineKeyboardButton("🧹 تصفية وفلترة متقدمة", callback_data="filter_file_data")
        ],
        [
            InlineKeyboardButton("❌ إلغاء القائمة", callback_data="cancel_file")
        ]
    ]

    await update.message.reply_text(panel_text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not query or not query.message:
        return
        
    await query.answer()
    chat_id = query.message.chat_id
    data = query.data

    if data == "mode_single":
        await query.message.reply_text("📝 يمكنك استخدام الأمر الفوري:\n`/dork <الاستعلام>`\nأو أرسل الدورك مباشرة في الرسالة القادمة.", parse_mode="Markdown")
        user_states[chat_id] = "WAITING_FOR_SINGLE_DORK"

    elif data == "mode_multi":
        await query.message.reply_text("📂 قم بإرسال ملف نصي `.txt` يحتوي على قائمة الدوركات مباشرة إلى المحادثة.")

    elif data == "start_file_process":
        dorks = user_files_data.get(chat_id, [])
        if not dorks:
            await query.message.reply_text("❌ لم يتم العثور على دوركات قيد الانتظار.")
            return

        total_dorks = len(dorks)
        await query.message.reply_text(f"🚀 *بدء معالجة {total_dorks} دورك عبر المحرك الخفي...*", parse_mode="Markdown")

        if chat_id not in dorker_instances or dorker_instances[chat_id] is None:
            dorker_instances[chat_id] = GoogleDorkerPro(headless=True, ultra_slow=False)

        dorker = dorker_instances[chat_id]
        loop = asyncio.get_running_loop()

        for i, dork in enumerate(dorks, 1):
            await loop.run_in_executor(None, dorker.search, dork)
            current_total_urls = len(dorker.found_urls)
            
            # إرسال تحديث بالتقدم الحالي بعد كل دورك
            await query.message.reply_text(
                f"🎯 *التقدم [{i}/{total_dorks}]*\n"
                f"🔍 **الدورك الحالي:** `{dork}`\n"
                f"🌐 **إجمالي الروابط الناتجة حتى الآن:** `{current_total_urls}`",
                parse_mode="Markdown"
            )
            
            if i < total_dorks:
                await asyncio.sleep(4)

        await query.message.reply_text(
            f"✅ *اكتمل فحص الملف بالكامل!*\n"
            f"📊 **إجمالي الدوركات:** `{total_dorks}`\n"
            f"🌐 **إجمالي الروابط الفريدة المستخرجة:** `{len(dorker.found_urls)}`",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )

    elif data == "filter_file_data":
        dorks = user_files_data.get(chat_id, [])
        filtered = [d for d in dorks if len(d) > 3 and not d.startswith("#")]
        user_files_data[chat_id] = filtered
        await query.message.reply_text(f"🧹 *تمت الفلترة بنجاح!*\nالعدد المتبقي بعد التنظيف: `{len(filtered)}`", parse_mode="Markdown")

    elif data == "cancel_file":
        user_files_data.pop(chat_id, None)
        await query.message.reply_text("❌ تم تفريغ القائمة المؤقتة.")

    elif data == "show_stats":
        await stats_command(update, context)

    elif data == "check_urls":
        if chat_id in dorker_instances and dorker_instances[chat_id]:
            await query.message.reply_text("🔍 *جاري التحقق من الروابط الشغالة (Alive / Dead)...*", parse_mode="Markdown")
            loop = asyncio.get_running_loop()
            alive, dead = await loop.run_in_executor(None, dorker_instances[chat_id].check_all_urls)
            await query.message.reply_text(f"✅ *اكتمل التحقق!*\n• الروابط الشغالة (Alive): `{len(alive)}`\n• الروابط الميتة (Dead): `{len(dead)}`", parse_mode="Markdown")
        else:
            await query.message.reply_text("❌ لا توجد جلسة عمل شغال حالياً.")

    elif data == "export_files":
        if chat_id in dorker_instances and dorker_instances[chat_id]:
            await query.message.reply_text("📦 *جاري إعداد الحزمة الكاملة وتصدير الملفات...*", parse_mode="Markdown")
            dorker = dorker_instances[chat_id]
            loop = asyncio.get_running_loop()
            txt, excel, pdf = await loop.run_in_executor(None, dorker.export_all_results)

            files_sent = 0
            for file_path in [txt, excel, pdf, dorker.output_file, dorker.emails_file]:
                # التثبت من وجود الملف وأن حجمه أكثر من 0 بايت لتجنب BadRequest Error
                if file_path and os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                    try:
                        with open(file_path, 'rb') as doc:
                            await context.bot.send_document(chat_id=chat_id, document=doc)
                        files_sent += 1
                    except Exception as e:
                        logger.error(f"فشل إرسال الملف {file_path}: {e}")
            
            if files_sent == 0:
                await query.message.reply_text("⚠️ لم يتم استخراج أي بيانات بعد، أو أن النتائج فارغة حالياً.")
        else:
            await query.message.reply_text("❌ لا توجد نتائج جارية للتصدير.")

    elif data == "stop_engine":
        if chat_id in dorker_instances and dorker_instances[chat_id]:
            await query.message.reply_text("🛑 *جاري حفظ الجلسة وإغلاق محرك Selenium...*", parse_mode="Markdown")
            dorker = dorker_instances[chat_id]
            loop = asyncio.get_running_loop()
            await loop.run_in_executor(None, dorker.close)
            dorker_instances[chat_id] = None
            await query.message.reply_text("👋 *تم إغلاق المحرك بنجاح وتأمين الحصيلة.*", parse_mode="Markdown")
        else:
            await query.message.reply_text("❌ المحرك متوقف بالفعل.")

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return
        
    chat_id = update.effective_chat.id
    state = user_states.get(chat_id)

    if state == "WAITING_FOR_SINGLE_DORK":
        dork = update.message.text.strip()
        user_states[chat_id] = None
        
        if chat_id not in dorker_instances or dorker_instances[chat_id] is None:
            dorker_instances[chat_id] = GoogleDorkerPro(headless=True, ultra_slow=False)

        dorker = dorker_instances[chat_id]
        await update.message.reply_text(f"🚀 *جاري تشغيل الفحص للدورك:*\n`{dork}`", parse_mode="Markdown")

        loop = asyncio.get_running_loop()
        added_now = await loop.run_in_executor(None, dorker.search, dork)
        await update.message.reply_text(
            f"✅ *اكتملت عملية الاستخراج!*\n"
            f"➕ روابط جديدة: `{added_now}`\n"
            f"🌐 المجموع الكلي للروابط: `{len(dorker.found_urls)}`",
            parse_mode="Markdown",
            reply_markup=get_main_keyboard()
        )
    else:
        await update.message.reply_text("يرجى اختيار أحد الأوامر من القائمة أو أرسل `/help`.", reply_markup=get_main_keyboard())

# ==============================================================================
# ERROR HANDLER AND INITIALIZATION
# ==============================================================================

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """معالج الأخطاء العالمي لتجنب توقف البوت أو انهياره"""
    logger.error("حدث استثناء أثناء تنفيذ طلب التلغرام:", exc_info=context.error)

async def post_init(application) -> None:
    """إزالة الـ Webhook وتصفية التحديثات المعلقة لمنع خطأ get_updates"""
    logger.info("جاري تنظيف وتصفية اتصالات Webhook المتبقية...")
    await application.bot.delete_webhook(drop_pending_updates=True)

# ==============================================================================
# MAIN BOT ENTRYPOINT
# ==============================================================================

if __name__ == "__main__":
    app = (
        ApplicationBuilder()
        .token(BOT_TOKEN)
        .post_init(post_init)
        .build()
    )

    # إضافة الأوامر
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("dork", dork_command))
    app.add_handler(CommandHandler("stats", stats_command))
    
    # إضافة معالجة الملفات والاستجابات
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))

    # تسجيل معالج الأخطاء
    app.add_error_handler(error_handler)

    print("🔥 [BOT RUNNING] البوت جاهز ويعمل بآلية التحكم الشاملة...")
    
    app.run_polling(
        drop_pending_updates=True,
        allowed_updates=Update.ALL_TYPES,
        poll_interval=1.0,
        timeout=30
    )
